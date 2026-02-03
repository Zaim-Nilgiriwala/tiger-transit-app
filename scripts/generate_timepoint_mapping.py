#!/usr/bin/env python3
"""
Generate a draft mapping from timepoint stop names (from the Timepoint Excel)
to GTFS stop IDs (from stops.json) using fuzzy matching.

Outputs:
  - data/processed/timepoint_mapping.json  (draft for user review)
  - Prints a review table to stdout

Usage:
  python scripts/generate_timepoint_mapping.py
"""

import json
import sys
from difflib import get_close_matches
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent

# --- Paths ---
EXCEL_PATH = PROJECT_ROOT / "mobile" / "src" / "ETA-Model" / "raw_data" / "Spring 2026 Timepoint Update.xlsx"
STOPS_JSON_PATH = PROJECT_ROOT / "mobile" / "src" / "ETA-Model" / "stops.json"
OUTPUT_PATH = PROJECT_ROOT / "data" / "processed" / "timepoint_mapping.json"


def load_stops(path: Path) -> dict:
    """Load stops.json and return {name: id} and {name_lower: name} dicts."""
    with open(path, "r", encoding="utf-8") as f:
        stops = json.load(f)
    name_to_id = {}
    for stop in stops:
        name_to_id[stop["name"]] = int(stop["id"])
    return name_to_id


def extract_timepoint_names(excel_path: Path) -> tuple:
    """
    Extract all unique timepoint stop names from row index 1 of each sheet.
    Returns:
        unique_names: set of unique timepoint names
        sheet_associations: dict mapping timepoint name -> list of sheet names
    """
    wb = openpyxl.load_workbook(str(excel_path), data_only=True)
    unique_names = set()
    sheet_associations = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue
        row1 = rows[1]
        for val in row1:
            if val is not None and isinstance(val, str) and val.strip():
                name = val.strip()
                unique_names.add(name)
                if name not in sheet_associations:
                    sheet_associations[name] = []
                if sheet_name not in sheet_associations[name]:
                    sheet_associations[name].append(sheet_name)

    wb.close()
    return unique_names, sheet_associations


def match_timepoint(name: str, name_to_id: dict) -> dict:
    """
    Attempt to match a timepoint name to a GTFS stop.
    Returns a dict with match details.
    """
    stop_names = list(name_to_id.keys())

    # 1. Exact match
    if name in name_to_id:
        return {
            "stop_id": name_to_id[name],
            "stop_name": name,
            "match_type": "exact",
            "confidence": "high",
            "alternatives": [],
        }

    # 2. Case-insensitive exact match
    name_lower = name.lower()
    for stop_name in stop_names:
        if stop_name.lower() == name_lower:
            return {
                "stop_id": name_to_id[stop_name],
                "stop_name": stop_name,
                "match_type": "case-insensitive",
                "confidence": "high",
                "alternatives": [],
            }

    # 3. Fuzzy match
    candidates = get_close_matches(name, stop_names, n=3, cutoff=0.5)
    if candidates:
        best = candidates[0]
        # Assign confidence based on similarity
        from difflib import SequenceMatcher
        ratio = SequenceMatcher(None, name.lower(), best.lower()).ratio()
        if ratio >= 0.85:
            confidence = "high"
        elif ratio >= 0.65:
            confidence = "medium"
        else:
            confidence = "low"

        return {
            "stop_id": name_to_id[best],
            "stop_name": best,
            "match_type": "fuzzy",
            "confidence": confidence,
            "alternatives": candidates,
        }

    # 4. No match
    # Try a broader fuzzy search with lower cutoff
    broad_candidates = get_close_matches(name, stop_names, n=3, cutoff=0.3)
    return {
        "stop_id": None,
        "stop_name": None,
        "match_type": "unmatched",
        "confidence": "none",
        "alternatives": broad_candidates,
    }


def print_review_table(mapping: dict, sheet_associations: dict) -> None:
    """Print a formatted review table to stdout."""
    print()
    print("=" * 80)
    print("  TIMEPOINT MAPPING DRAFT - REVIEW REQUIRED")
    print("=" * 80)
    print()

    # Group by match type for easier review
    exact = []
    case_insensitive = []
    fuzzy = []
    unmatched = []

    for name, info in sorted(mapping.items()):
        entry = (name, info)
        if info["match_type"] == "exact":
            exact.append(entry)
        elif info["match_type"] == "case-insensitive":
            case_insensitive.append(entry)
        elif info["match_type"] == "fuzzy":
            fuzzy.append(entry)
        else:
            unmatched.append(entry)

    # Print exact matches
    if exact:
        print(f"--- EXACT MATCHES ({len(exact)}) ---")
        for name, info in exact:
            routes = ", ".join(sheet_associations.get(name, []))
            print(f'  [EXACT]  "{name}" -> "{info["stop_name"]}" (stop_id: {info["stop_id"]})')
            print(f'           Routes: {routes}')
        print()

    # Print case-insensitive matches
    if case_insensitive:
        print(f"--- CASE-INSENSITIVE MATCHES ({len(case_insensitive)}) ---")
        for name, info in case_insensitive:
            routes = ", ".join(sheet_associations.get(name, []))
            print(f'  [CASE]   "{name}" -> "{info["stop_name"]}" (stop_id: {info["stop_id"]})')
            print(f'           Routes: {routes}')
        print()

    # Print fuzzy matches (NEED REVIEW)
    if fuzzy:
        print(f"--- FUZZY MATCHES ({len(fuzzy)}) - VERIFY THESE ---")
        for name, info in fuzzy:
            routes = ", ".join(sheet_associations.get(name, []))
            alts = ", ".join(f'"{a}"' for a in info["alternatives"])
            print(f'  [FUZZY]  "{name}" -> "{info["stop_name"]}" (stop_id: {info["stop_id"]}) -- confidence: {info["confidence"]}')
            print(f'           Candidates: [{alts}]')
            print(f'           Routes: {routes}')
        print()

    # Print unmatched (NEED USER INPUT)
    if unmatched:
        print(f"--- UNMATCHED ({len(unmatched)}) - NEED MANUAL MAPPING ---")
        for name, info in unmatched:
            routes = ", ".join(sheet_associations.get(name, []))
            alts = ", ".join(f'"{a}"' for a in info["alternatives"])
            print(f'  [NONE]   "{name}" -> ??? ')
            print(f'           Candidates: [{alts}]')
            print(f'           Routes: {routes}')
        print()

    # Summary
    total = len(mapping)
    matched = sum(1 for v in mapping.values() if v["stop_id"] is not None)
    print("=" * 80)
    print(f"  SUMMARY: {matched}/{total} matched, {total - matched} need manual review")
    print(f"  Output: {OUTPUT_PATH}")
    print("=" * 80)
    print()


def main():
    # Validate paths
    if not EXCEL_PATH.exists():
        print(f"ERROR: Timepoint Excel not found at {EXCEL_PATH}")
        sys.exit(1)
    if not STOPS_JSON_PATH.exists():
        print(f"ERROR: stops.json not found at {STOPS_JSON_PATH}")
        sys.exit(1)

    # Load GTFS stops
    print(f"Loading stops from {STOPS_JSON_PATH}...")
    name_to_id = load_stops(STOPS_JSON_PATH)
    print(f"  Loaded {len(name_to_id)} GTFS stops")

    # Extract timepoint names from Excel
    print(f"Extracting timepoint names from {EXCEL_PATH}...")
    unique_names, sheet_associations = extract_timepoint_names(EXCEL_PATH)
    print(f"  Found {len(unique_names)} unique timepoint names across {len(sheet_associations)} associations")

    # Match each timepoint name
    print("Matching timepoint names to GTFS stops...")
    mapping = {}
    for name in sorted(unique_names):
        mapping[name] = match_timepoint(name, name_to_id)

    # Write draft mapping
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(mapping, f, indent=2, ensure_ascii=False)
    print(f"Draft mapping written to {OUTPUT_PATH}")

    # Print review table
    print_review_table(mapping, sheet_associations)


if __name__ == "__main__":
    main()
