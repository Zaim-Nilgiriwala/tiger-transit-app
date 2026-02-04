#!/usr/bin/env python3
"""
Parse timepoint schedule data from the Timepoint Excel workbook and the
verified timepoint-to-GTFS mapping into a structured parquet file.

Requires:
  - data/processed/timepoint_mapping.json  (reviewed by user)
  - mobile/src/ETA-Model/raw_data/Spring 2026 Timepoint Update.xlsx
  - gtfs_data/routes.txt

Outputs:
  - data/processed/timepoints.parquet

Usage:
  python scripts/parse_timepoints.py
"""

import csv
import datetime
import json
import sys
from difflib import SequenceMatcher, get_close_matches
from pathlib import Path

import pandas as pd
import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent

# --- Paths ---
MAPPING_PATH = PROJECT_ROOT / "data" / "processed" / "timepoint_mapping.json"
RAW_DATA_DIR = PROJECT_ROOT / "mobile" / "src" / "ETA-Model" / "raw_data"
ROUTES_PATH = PROJECT_ROOT / "gtfs_data" / "routes.txt"
OUTPUT_PATH = PROJECT_ROOT / "data" / "processed" / "timepoints.parquet"


def find_excel_file(directory: Path) -> Path:
    """Find the timepoint Excel file in the raw data directory."""
    candidates = list(directory.glob("*Timepoint*.xlsx")) + list(
        directory.glob("*timepoint*.xlsx")
    )
    if not candidates:
        # Fall back to any xlsx file
        candidates = list(directory.glob("*.xlsx"))
    if not candidates:
        print(f"ERROR: No .xlsx file found in {directory}")
        sys.exit(1)
    if len(candidates) > 1:
        print(f"WARNING: Multiple .xlsx files found, using: {candidates[0].name}")
    return candidates[0]


def load_mapping(path: Path) -> dict:
    """Load the verified timepoint mapping. Returns {name: stop_id_or_None}."""
    if not path.exists():
        print(
            f"ERROR: Mapping file not found at {path}\n"
            "Run 'python scripts/generate_timepoint_mapping.py' first, then review the output."
        )
        sys.exit(1)
    with open(path, "r", encoding="utf-8") as f:
        raw = json.load(f)
    mapping = {}
    for name, info in raw.items():
        mapping[name] = info.get("stop_id")  # None for unmatched
    return mapping


def load_route_mapping(path: Path) -> dict:
    """Load routes.txt and build route_long_name -> route_id mapping."""
    if not path.exists():
        print(f"ERROR: routes.txt not found at {path}")
        sys.exit(1)
    name_to_id = {}
    with open(path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            route_id_raw = row["route_id"].strip().strip('"')
            long_name = row["route_long_name"].strip().strip('"')
            # Extract first numeric segment as route_id
            # e.g. "215_202_201_156" -> 215, "235_93" -> 235
            parts = route_id_raw.split("_")
            try:
                route_id_num = int(parts[0])
            except ValueError:
                continue
            name_to_id[long_name] = route_id_num
    return name_to_id


# Manual overrides: sheet name -> route long name (from routes.txt)
# Handles cases where sheet names don't exactly match route_long_name
SHEET_TO_ROUTE_NAME = {
    "Glenn Harper": "Glenn-Harper",
    "Old Row": "Old Row - West Parking",
    "Samford Shug": "Samford-Shug Jordan",
    "North Dean": "North Dean -AM  (before 11 am)",  # Map to AM variant
    "SQ-FA": "South Quad/Fine Arts",
    "South Quad": "South Quad/Fine Arts",  # Same route, different sub-table
    "P&R": "Park & Ride",
    "Heath Science": "Health Sciences Sector",
}


def match_sheet_to_route(sheet_name: str, route_names: dict) -> tuple:
    """
    Match a sheet name to a route in routes.txt.
    Returns (route_id, route_long_name) or (None, None).
    """
    # 1. Check manual overrides
    if sheet_name in SHEET_TO_ROUTE_NAME:
        mapped_name = SHEET_TO_ROUTE_NAME[sheet_name]
        if mapped_name in route_names:
            return route_names[mapped_name], mapped_name

    # 2. Exact match
    if sheet_name in route_names:
        return route_names[sheet_name], sheet_name

    # 3. Case-insensitive match
    for rname in route_names:
        if rname.lower() == sheet_name.lower():
            return route_names[rname], rname

    # 4. Fuzzy match
    candidates = get_close_matches(sheet_name, list(route_names.keys()), n=1, cutoff=0.5)
    if candidates:
        best = candidates[0]
        return route_names[best], best

    return None, None


def detect_column_groups(header_row: tuple) -> list:
    """
    Detect groups of consecutive non-None columns in the header row.
    Each group is a list of (col_index, stop_name) tuples.
    Groups are separated by None/blank columns.
    """
    groups = []
    current_group = []

    for i, val in enumerate(header_row):
        if val is not None and isinstance(val, str) and val.strip():
            current_group.append((i, val.strip()))
        else:
            if current_group:
                groups.append(current_group)
                current_group = []
    if current_group:
        groups.append(current_group)

    return groups


def parse_sheet(ws, sheet_name: str, timepoint_mapping: dict, route_id: int) -> list:
    """
    Parse a single worksheet into schedule tuples.
    Returns list of dicts with route_id, route_name, stop_id, stop_name, scheduled_time.
    """
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        print(f"  WARNING: Sheet '{sheet_name}' has fewer than 3 rows, skipping")
        return []

    # Row 0: description (ignored for parsing)
    # Row 1: timepoint stop names
    # Row 2+: schedule times
    header_row = rows[1]
    groups = detect_column_groups(header_row)

    if not groups:
        print(f"  WARNING: No column groups found in sheet '{sheet_name}'")
        return []

    records = []
    skipped_stops = set()

    for group in groups:
        for col_idx, stop_name in group:
            # Look up stop_id from mapping
            stop_id = timepoint_mapping.get(stop_name)
            if stop_id is None:
                if stop_name not in skipped_stops:
                    skipped_stops.add(stop_name)
                    if stop_name in timepoint_mapping:
                        print(
                            f"  WARNING: Skipping unmatched timepoint '{stop_name}' "
                            f"in sheet '{sheet_name}' (no GTFS stop_id)"
                        )
                    else:
                        print(
                            f"  WARNING: Timepoint '{stop_name}' in sheet '{sheet_name}' "
                            f"not found in mapping file"
                        )
                continue

            # Parse time rows
            for row_idx in range(2, len(rows)):
                row = rows[row_idx]
                if col_idx >= len(row):
                    continue
                cell_val = row[col_idx]
                if cell_val is None:
                    continue
                if isinstance(cell_val, datetime.time):
                    time_str = cell_val.strftime("%H:%M:%S")
                    records.append(
                        {
                            "route_id": route_id,
                            "route_name": sheet_name,
                            "stop_id": stop_id,
                            "stop_name": stop_name,
                            "scheduled_time": time_str,
                        }
                    )
                elif isinstance(cell_val, datetime.datetime):
                    # Sometimes Excel stores times as datetime
                    time_str = cell_val.strftime("%H:%M:%S")
                    records.append(
                        {
                            "route_id": route_id,
                            "route_name": sheet_name,
                            "stop_id": stop_id,
                            "stop_name": stop_name,
                            "scheduled_time": time_str,
                        }
                    )

    return records


def main():
    print("=" * 60)
    print("  TIMEPOINT PARSER")
    print("=" * 60)

    # Load verified mapping
    print(f"\nLoading verified mapping from {MAPPING_PATH}...")
    timepoint_mapping = load_mapping(MAPPING_PATH)
    matched = sum(1 for v in timepoint_mapping.values() if v is not None)
    unmatched = sum(1 for v in timepoint_mapping.values() if v is None)
    print(f"  {matched} matched, {unmatched} unmatched (will be skipped)")

    # Find Excel file
    excel_path = find_excel_file(RAW_DATA_DIR)
    print(f"\nLoading workbook: {excel_path.name}...")
    wb = openpyxl.load_workbook(str(excel_path), data_only=True)
    print(f"  Found {len(wb.sheetnames)} sheets: {', '.join(wb.sheetnames)}")

    # Load route mapping
    print(f"\nLoading routes from {ROUTES_PATH}...")
    route_name_to_id = load_route_mapping(ROUTES_PATH)
    print(f"  Loaded {len(route_name_to_id)} routes")

    # Parse each sheet
    print("\nParsing sheets...")
    all_records = []
    sheets_processed = 0
    sheets_skipped = []
    route_counts = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        route_id, route_long_name = match_sheet_to_route(sheet_name, route_name_to_id)

        if route_id is None:
            print(f"  SKIP: No route match for sheet '{sheet_name}'")
            sheets_skipped.append(sheet_name)
            continue

        records = parse_sheet(ws, sheet_name, timepoint_mapping, route_id)
        all_records.extend(records)
        sheets_processed += 1

        if records:
            route_counts[sheet_name] = len(records)
            print(
                f"  OK: '{sheet_name}' -> route {route_id} ({route_long_name}) "
                f"-- {len(records)} time entries"
            )
        else:
            print(f"  EMPTY: '{sheet_name}' -> route {route_id} -- 0 time entries")

    wb.close()

    if not all_records:
        print("\nERROR: No records parsed from any sheet!")
        sys.exit(1)

    # Build DataFrame
    df = pd.DataFrame(all_records)
    df["route_id"] = df["route_id"].astype(int)
    df["stop_id"] = df["stop_id"].astype(int)

    # Write output
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    df.to_parquet(OUTPUT_PATH, index=False)

    # Validation summary
    print("\n" + "=" * 60)
    print("  VALIDATION SUMMARY")
    print("=" * 60)
    print(f"\n  Total time entries:  {len(df):,}")
    print(f"  Unique routes:       {df['route_id'].nunique()}")
    print(f"  Unique stops:        {df['stop_id'].nunique()}")
    print(f"  Sheets processed:    {sheets_processed}/{len(wb.sheetnames)}")
    if sheets_skipped:
        print(f"  Sheets skipped:      {', '.join(sheets_skipped)}")

    print("\n  Entries per route:")
    for sheet_name in sorted(route_counts, key=lambda x: route_counts[x], reverse=True):
        print(f"    {sheet_name:<25} {route_counts[sheet_name]:>5}")

    print(f"\n  Output: {OUTPUT_PATH}")
    print(f"  Size:   {OUTPUT_PATH.stat().st_size / 1024:.1f} KB")
    print("=" * 60)


if __name__ == "__main__":
    main()
